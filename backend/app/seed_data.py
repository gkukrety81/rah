import os, math, re, asyncio
from typing import Optional, Dict, Any, List

from openpyxl import load_workbook
from pypdf import PdfReader

from sqlalchemy import select, update
from sqlalchemy.dialects.postgresql import insert as pg_insert

from .db import engine, SessionLocal
from .models import PhysiologyProgram, RahItem, RahItemProgram

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
PDF_PATTERN = re.compile(r"^\s*(\d{1,3})\.00\s+(.+?)\s*$")

def _floor_program_code(rah_id: float) -> int:
    return int(math.floor(rah_id))

def _as_float(x: Any) -> Optional[float]:
    try:
        if x is None or (isinstance(x, str) and not x.strip()):
            return None
        return float(x)
    except Exception:
        return None

def _clean_str(x: Any) -> Optional[str]:
    if x is None:
        return None
    s = str(x).strip()
    return s if s else None

async def upsert_program(session, program_code: int, name: str, sex: str = "unisex"):
    stmt = pg_insert(PhysiologyProgram).values(
        program_code=program_code, name=name, sex=sex
    ).on_conflict_do_update(
        index_elements=[PhysiologyProgram.program_code],
        set_={"name": name, "sex": sex}
    )
    await session.execute(stmt)

async def upsert_rah_item(session, rah_id: float, details: Optional[str], category: Optional[str]):
    stmt = pg_insert(RahItem).values(
        rah_id=rah_id, details=details, category=category
    ).on_conflict_do_update(
        index_elements=[RahItem.rah_id],
        set_={"details": details, "category": category}
    )
    await session.execute(stmt)

async def ensure_mapping(session, rah_id: float, program_code: int):
    stmt = pg_insert(RahItemProgram).values(
        rah_id=rah_id, program_code=program_code
    ).on_conflict_do_nothing()
    await session.execute(stmt)

def parse_programs_from_pdfs() -> List[Dict[str, Any]]:
    """
    Scans PDFs in DATA_DIR and extracts lines like:
    '58.00 Acoustic organ / organ of equilibrium, phy'
    Returns list of dicts: [{'program_code': 58, 'name': 'Acoustic organ ...'}]
    """
    programs: Dict[int, str] = {}
    if not os.path.isdir(DATA_DIR):
        return []

    for fname in os.listdir(DATA_DIR):
        if not fname.lower().endswith(".pdf"):
            continue
        path = os.path.join(DATA_DIR, fname)
        try:
            reader = PdfReader(path)
            for page in reader.pages:
                text = page.extract_text() or ""
                for line in text.splitlines():
                    m = PDF_PATTERN.match(line)
                    if m:
                        code = int(m.group(1))
                        name = m.group(2).strip()
                        # keep first seen, or update if longer (heuristic)
                        if code not in programs or len(name) > len(programs[code]):
                            programs[code] = name
        except Exception:
            # ignore unreadable PDFs; we can add programs later via UI
            pass

    return [{"program_code": k, "name": v, "sex": "unisex"} for k, v in sorted(programs.items())]

def load_excel_rows(xls_path: str) -> List[Dict[str, Any]]:
    """
    Reads 'RAH List.xlsx' and returns rows with keys:
    rah_id, details, category, description (if present), correlation (if present)
    """
    wb = load_workbook(xls_path, data_only=True)
    ws = wb.active

    # Normalize header map
    headers = {}
    for i, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1, values_only=True))):
        if cell is None:
            continue
        name = str(cell).strip().lower()
        headers[name] = i

    def col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        return None

    col_rah_id = col("rah id", "rah_id", "id")
    col_details = col("details", "name", "item")
    col_category = col("category", "group")
    col_description = col("description", "desc")
    col_correlation = col("correlation", "correl", "corr")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rah_id = _as_float(r[col_rah_id]) if col_rah_id is not None else None
        if rah_id is None:
            continue
        rows.append({
            "rah_id": rah_id,
            "details": _clean_str(r[col_details]) if col_details is not None else None,
            "category": _clean_str(r[col_category]) if col_category is not None else None,
            "description": _clean_str(r[col_description]) if col_description is not None else None,
            "correlation": _clean_str(r[col_correlation]) if col_correlation is not None else None,
        })
    return rows

async def seed():
    # 1) Upsert programs (from PDFs if available)
    pdf_programs = parse_programs_from_pdfs()

    async with SessionLocal() as session:
        # If we found programs in PDFs, upsert them
        if pdf_programs:
            for p in pdf_programs:
                await upsert_program(session, p["program_code"], p["name"], p["sex"])
        await session.commit()

        # 2) Load RAH List.xlsx if present
        xls_path = os.path.join(DATA_DIR, "RAH List.xlsx")
        if os.path.isfile(xls_path):
            rows = load_excel_rows(xls_path)

            # Upsert rah_item and associations by integer part → program_code
            for row in rows:
                rah_id = row["rah_id"]
                details = row["details"]
                category = row["category"]
                await upsert_rah_item(session, rah_id, details, category)

                program_code = _floor_program_code(rah_id)
                # If program missing, create a placeholder; can be edited later
                await upsert_program(session, program_code, f"Program {program_code}.00", "unisex")
                await ensure_mapping(session, rah_id, program_code)

            await session.commit()

    print("Seeding complete.")
    if pdf_programs:
        print(f"Upserted programs from PDFs: {len(pdf_programs)}")
    if os.path.isfile(os.path.join(DATA_DIR, 'RAH List.xlsx')):
        print("Imported RAH items from Excel and mapped to programs.")

if __name__ == "__main__":
    asyncio.run(seed())
